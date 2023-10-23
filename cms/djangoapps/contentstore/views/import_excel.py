
from common.djangoapps.edxmako.shortcuts import render_to_response
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import ensure_csrf_cookie
from django.core.exceptions import PermissionDenied
from cms.djangoapps.contentstore.views.certificates import _get_course_and_check_access
from common.djangoapps.util.json_request import JsonResponse
from opaque_keys.edx.keys import  CourseKey
from common.djangoapps.edxmako.shortcuts import render_to_response
from cms.djangoapps.contentstore.views.item import create_item_import , _update_with_callback ,_save_xblock 
# from cms.djangoapps.contentstore.models import CourseUnitTime
from openedx.core.djangoapps.content.course_overviews.models import CourseUnitTime
import openpyxl

@login_required
@ensure_csrf_cookie
def viewImportExcel (request, course_id) :
    course_key = CourseKey.from_string(course_id)
   
    try:
        course = _get_course_and_check_access(course_key, request.user)
    except PermissionDenied:
        msg = _('PermissionDenied: Failed in authenticating {user}').format(user=request.user)
        return JsonResponse({"error": msg}, status=403)
    context = {
          'context_course': course,
    }
    parts = course_id.split('+')
    organization = parts[0].split(":")[1]  
    course_number = parts[1]
    course_run = parts[2]
    parent_locator = f'block-v1:{organization}+{course_number}+{course_run}+type@course+block@course'
    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]
       
        if excel_file.name.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            sheets = workbook.sheetnames
            Details = workbook['2. Details']       
            Duration = workbook['3. Duration']
            section=create_item_import(request, parent_locator, category='chapter', display_name='Mở đầu')

            section_block = create_item_import(request ,parent_locator, category='chapter', display_name='Nội dung khoá học')
           
            # sheet details
            list_video = []
            url_ = ''
            list_unit = []
            for data in Details.iter_rows(values_only=True):
                
                if data[1] is not None :
                    if ':' in data[1] :
                   
                        lesson_block =create_item_import(request, parent_locator=str(section_block.location) , category='sequential', display_name=data[1]) 
                        create_item_import(request, parent_locator=str(lesson_block.location), category='vertical', display_name='Mở đầu')
                        unit_block=create_item_import(request, parent_locator=str(lesson_block.location), category='vertical', display_name='Nội dung bài học')
                        html_block =create_item_import(request, parent_locator= str(unit_block.location) ,category='html')  
                        list_unit.append({'unit' : data[1] , "block_id" : str(lesson_block.location) })
                    if data[3] is not None and 'video' in data[3].lower() and data[5] is not None:
                        videos =[]
                        videos.append(data[5])
                        list_video.append({'title' : data[2], 'video' : videos , 'unit' : data[1] , 'block': html_block})

            newArr = []  
            grouped = {} 
            for item in list_video:
                block_value = item['block']
                if block_value not in grouped:
                    grouped[block_value] = []  
                grouped[block_value].append({'title': item['title'], 'video' : item['video']})  #   
                     
           
            for block_value, content in grouped.items():
                newArr.append({'block' : block_value , 'content' : content})

            for block in newArr :
                url_ =''
                for a in block['content']:
                    for v in a['video']:
                        url_ += f'<p><a href="{v}" target="_blank" title={a["title"]}>Video: {a["title"]}</a></p>'
                _save_xblock(xblock=block['block'] , user=request.user, data=url_)

            # sheet Duration
            for data in Duration.iter_rows(values_only=True):
                if data[0] is not None:
                    for u in list_unit :
                        if data[0] in u['unit']:
                          
                            CourseUnitTime.create_unit_time(course_id=course_id, block_id=u['block_id'], display_name=u['unit'], total=data[10])
            
            workbook.close()
           
                  
    return render_to_response('import_excel.html', context)